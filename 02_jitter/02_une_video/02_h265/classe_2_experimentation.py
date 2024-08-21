import os
import re
from mininet.topo import Topo
from mininet.net import Mininet
from mininet.node import RemoteController, OVSKernelSwitch
from mininet.cli import dumpNodeConnections, CLI
from mininet.log import setLogLevel
from mininet.link import TCLink
import time
import subprocess
import threading
import shutil
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from typing import Dict, List, Tuple
from openpyxl import Workbook


class VideoInput:
    def __init__(self):
        self._servers = ['h1']
        self._clients = ['h2']
        self._HOST_NUMBERS = [2]
        #self._PERTURBATION_NUMBERS = [16, 15, 12, 10, 8, 6, 4, 3]
        self._PERTURBATION_NUMBERS = [20, 50, 100, 200, 500, 1000]  # Configurations delay en (en ms)
        self._JITTERS_NUMBERS = [2, 5, 10, 20, 50, 100]  # Configurations jitters en (en ms)
        self._ports = [5000]        
        self._nginx_configs = ["/usr/local/nginx/conf/sdn/nginxh1.conf"]
        self._video_urls = ["http://10.1.1.1:5000/h265/hls/1080/240/bbb_h265_slave_240_4.m3u8"]
        self._nom_interface_attendus = ["h2-eth0"]
        self._protocole = "hls"
        self._codec = "h265"
        self._perturbation = "delay"
        self._nombre_hotes = 2
        self._profile = 240

    def get_servers(self):
        return self._servers

    def get_clients(self):
        return self._clients

    def get_host_numbers(self):
        return self._HOST_NUMBERS

    def get_perturbation_numbers(self):
        return self._PERTURBATION_NUMBERS
        
    def get_jitters_numbers(self):
        return self._JITTERS_NUMBERS

    def get_ports(self):
        return self._ports

    def get_nginx_configs(self):
        return self._nginx_configs

    def get_video_urls(self):
        return self._video_urls

    def get_nom_interface_attendus(self):
        return self._nom_interface_attendus

    def get_protocole(self):
        return self._protocole

    def get_codec(self):
        return self._codec

    def get_perturbation(self):
        return self._perturbation

    def get_nombre_hotes(self):
        return self._nombre_hotes

    def get_profile(self):
        return self._profile
        

class SetupScript:
    def setup_controller_connection(self):
        topo = NetworkTopology()
        net = Mininet(topo=topo, controller=lambda name: RemoteController(name, ip='127.0.0.1', port=6633), switch=OVSKernelSwitch, link=TCLink)
        return net

    def setup_nginx_server(self):
        # Implement nginx setup here
        pass

class NetworkTopology(Topo):
    def build(self):
        s1 = self.addSwitch('s1')
        s2 = self.addSwitch('s2')
        for i in range(1, 17):
            self.addHost(f'h{i}', mac=f"00:00:00:00:00:{i:02X}", ip=f"10.1.1.{i}/24")
        self.addLink(s1, s2, cls=TCLink, bw=16)
        for i in range(1, 17, 2):
            self.addLink(f'h{i}', s1, cls=TCLink, bw=16)
            self.addLink(f'h{i+1}', s2, cls=TCLink, bw=16)

    def config_link_status(self, s1, s2, status, bw):
        # Code to configure the link status between s1 and s2
        pass

class VideoStreamingClient:
    @staticmethod
    def start_streaming(server, client, port, nginx_config, video_url, nom_interface_attendu, protocole, codec, perturbation, nombre_hotes, bw):
        nom_repertoire = f"chunks-{server}_{client}_bbb_{codec}_{protocole}_hotes_{nombre_hotes}_{perturbation}_{bw}"
        nom_fichier = f"{server}_{client}_bbb_{codec}_{protocole}_hotes_{nombre_hotes}_{perturbation}_{bw}.pcapng"
        if not os.path.exists(nom_repertoire):
            try:
                os.mkdir(nom_repertoire)
            except OSError:
                pass
        if not os.path.exists(nom_fichier):
            try:
                open(nom_fichier, 'w').close()
            except OSError:
                pass
        server.cmd(f'nginx -c {nginx_config}')
        fichier_sortie = os.path.join(nom_repertoire, "chunk_%03d.ts")
        client.cmd(f'ffmpeg -i "{video_url}" -c copy -f segment -segment_time 2 -segment_wrap 32 "{fichier_sortie}" &')
        client.cmd(f'tshark -i "{nom_interface_attendu}" -w "{nom_fichier}" &')
        client.cmd(f'ffplay -autoexit {video_url}')
        time.sleep(30)
        print(f"Les chunks ont été téléchargés avec succès dans {nom_repertoire}.")
        print(f"La capture a été arrêtée et le fichier {nom_fichier} est maintenant fermé.")

class QoSMetricsCollector:
    def __init__(self, data_pcapng_files: Dict[str, Dict[int, str]], bash_script_path: str):
        self.data_pcapng_files = data_pcapng_files
        self.bash_script_path = bash_script_path
        self.data_qos_files_bitrate = {}
        self.data_qos_files_packet_loss = {}
        self.data_qos_files_average_latency = {}
        self.data_qos_files_average_jitter = {}
        self.lock = threading.Lock()

    def _get_ip_address(self, key: str, is_server: bool) -> str:
        hX, hY = key.split('_')
        if is_server:
            ip_num = hX[1:]
        else:
            ip_num = hY[1:]
        return f"10.1.1.{ip_num}"

    def calculate_metrics_for_pcapng(self, file_path, ip_src, ip_dst) -> Tuple[float, float, float, float]:
        bash_command = f"bash {self.bash_script_path} {file_path} {ip_src} {ip_dst}"
        result = subprocess.run(bash_command, shell=True, capture_output=True, text=True)
        output = result.stdout.strip()
        pattern = r"bitrate: (\d+[\.,]?\d*) packet_loss: (\d+[\.,]?\d*) average_latency: (\d+[\.,]?\d*) average_jitter: (\d+[\.,]?\d*)"
        match = re.search(pattern, output)
        if not match:
            print("Erreur: Impossible de trouver les métriques dans la sortie du script bash.")
            return (0.0, 0.0, 0.0, 0.0)
        return (float(match.group(1).replace(',', '.')),
                float(match.group(2).replace(',', '.')),
                float(match.group(3).replace(',', '.')),
                float(match.group(4).replace(',', '.')))

    def calculate_qos_metrics(self, key):
        ip_src = self._get_ip_address(key, True)
        ip_dst = self._get_ip_address(key, False)
        resolutions = sorted(self.data_pcapng_files[key].keys())
        for resolution in resolutions:
            file_path = self.data_pcapng_files[key][resolution]
            metrics = self.calculate_metrics_for_pcapng(file_path, ip_src, ip_dst)
            with self.lock:
                if key not in self.data_qos_files_bitrate:
                    self.data_qos_files_bitrate[key] = {}
                    self.data_qos_files_packet_loss[key] = {}
                    self.data_qos_files_average_latency[key] = {}
                    self.data_qos_files_average_jitter[key] = {}
                self.data_qos_files_bitrate[key][resolution] = metrics[0]
                self.data_qos_files_packet_loss[key][resolution] = metrics[1]
                self.data_qos_files_average_latency[key][resolution] = metrics[2]
                self.data_qos_files_average_jitter[key][resolution] = metrics[3]

    def run(self):
        threads = []
        for key in list(self.data_pcapng_files.keys())[:4]:
            thread = threading.Thread(target=self.calculate_qos_metrics, args=(key,))
            threads.append(thread)
            thread.start()
        for thread in threads:
            thread.join()
        return (self.data_qos_files_bitrate, self.data_qos_files_packet_loss, self.data_qos_files_average_latency, self.data_qos_files_average_jitter)

class QoEMetricsCollector:
    def __init__(self, data_video_files, chemin_video_serveur):
        self.data_video_files = data_video_files
        self.chemin_video_serveur = chemin_video_serveur
        self.data_video_files_psnr = {}
        self.data_video_files_ssim = {}
        self.lock = threading.Lock()

    def calculate_metrics(self, key, videos):
        psnr_results = {}
        ssim_results = {}
        for resolution, video_path in videos.items():
            psnr_results[resolution] = self.calcul_psnr(self.chemin_video_serveur, video_path)
            ssim_results[resolution] = self.calcul_ssim(self.chemin_video_serveur, video_path)
        with self.lock:
            self.data_video_files_psnr[key] = psnr_results
            self.data_video_files_ssim[key] = ssim_results

    def run(self):
        threads = []
        for key, videos in list(self.data_video_files.items())[:4]:
            thread = threading.Thread(target=self.calculate_metrics, args=(key, videos))
            threads.append(thread)
            thread.start()
        for thread in threads:
            thread.join()
        return self.data_video_files_psnr, self.data_video_files_ssim

    def calcul_psnr(self, video_path_1, video_path_2) -> float:
        psnr_process = subprocess.run(["ffmpeg", "-i", video_path_1, "-i", video_path_2, "-lavfi", "psnr", "-f", "null", "-"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        psnr_output = psnr_process.stderr.decode()
        psnr_match = re.search(r'average:(\d+\.\d+)', psnr_output)
        if psnr_match:
            return float(psnr_match.group(1))
        else:
            raise ValueError("Pas de valeur PSNR trouvée.")

    def calcul_ssim(self, video_path_1, video_path_2) -> float:
        ssim_process = subprocess.run(["ffmpeg", "-i", video_path_1, "-i", video_path_2, "-lavfi", "ssim", "-f", "null", "-"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        ssim_output = ssim_process.stderr.decode()
        ssim_match = re.search(r'All:(\d+\.\d+)', ssim_output)
        if ssim_match:
            return float(ssim_match.group(1))
        else:
            raise ValueError("Pas de valeur SSIM trouvée.")

class ExcelSaver:
    @staticmethod
    def save_metrics_to_excel(data_dict, dict_name, codec, profile, perturbation, protocole, nombre_hotes):
        from_dict_name = dict_name.split('_')[-1]
        excel_files_dir = f'excel_files_{codec}_{profile}_{from_dict_name}'
        os.makedirs(excel_files_dir, exist_ok=True)
        excel_file_path = os.path.join(excel_files_dir, f'resultats_{perturbation}_{protocole}_hotes_{nombre_hotes}_{codec}_{profile}_{from_dict_name}.xlsx')
        workbook = Workbook()
        workbook.save(excel_file_path)
        images_dir = os.path.join(excel_files_dir, f'images_{codec}_{profile}_{from_dict_name}')
        os.makedirs(images_dir, exist_ok=True)
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active
        headers = [f'{perturbation}'] + list(data_dict.keys())
        sheet.append(headers)
        all_keys = sorted(set().union(*(d.keys() for d in data_dict.values())))
        for key in all_keys:
            row = [key] + [data_dict[column].get(key, '') for column in headers[1:]]
            sheet.append(row)
        wb.save(excel_file_path)
        for index, (key, value_dict) in enumerate(data_dict.items(), start=1):
            sheet_name = f'Graph_{key}'
            wb.create_sheet(title=sheet_name)
            new_sheet = wb[sheet_name]
            plt.figure()
            lists = sorted(value_dict.items())
            x, y = zip(*lists)
            plt.plot(x, y, marker='o')
            plt.title(f'{perturbation} {protocole} avec {nombre_hotes} hotes codec {codec} profile {profile} : {key}')
            plt.xlabel(f'{perturbation}')
            plt.ylabel(f'{from_dict_name}')
            image_path = os.path.join(images_dir, f'plot_{key}.png')
            plt.savefig(image_path)
            plt.close()
            img = Image(image_path)
            new_sheet.add_image(img, 'A1')
        wb.save(excel_file_path)

#########

class DataOrganizer:
    @staticmethod
    def createDataExperimentDirectory(move_data_bash_script_path):
        # Préparer la commande bash
        bash_command = f"bash {move_data_bash_script_path}"
        
        # Exécuter la commande bash
        process = subprocess.Popen(bash_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, stderr = process.communicate()
        
        if process.returncode == 0:
            print(stdout.decode())
        else:
            print(stderr.decode())


    @staticmethod
    def transfertContentDataExperimentToRepertoireData(move_data_to_repertoire_data_bash_script_path, *args):
        # Préparer la commande bash avec les arguments
        bash_command = f"bash {move_data_to_repertoire_data_bash_script_path} {' '.join(args)}"
        
        # Exécuter la commande bash
        process = subprocess.Popen(bash_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, stderr = process.communicate()
        
        if process.returncode == 0:
            print(stdout.decode())
        else:
            print(stderr.decode())

    @staticmethod
    def read_data_from_files():
        data_directories = {}
        data_files = {}

        with open('directories.txt', 'r') as f:
            for line in f:
                server_client_key, perturbation_number, target_dir = line.strip().split(',')
                if server_client_key not in data_directories:
                    data_directories[server_client_key] = {}
                data_directories[server_client_key][int(perturbation_number)] = target_dir

        with open('files.txt', 'r') as f:
            for line in f:
                server_client_key, perturbation_number, source_file = line.strip().split(',')
                if server_client_key not in data_files:
                    data_files[server_client_key] = {}
                data_files[server_client_key][int(perturbation_number)] = source_file

        return data_directories, data_files

    @staticmethod
    def read_data_from_files2():
        data_directories = {}
        data_files = {}

        with open('directories.txt', 'r') as f:
            for line in f:
                server_client_key, perturbation_string, target_dir = line.strip().split(',')
                if server_client_key not in data_directories:
                    data_directories[server_client_key] = {}
                data_directories[server_client_key][perturbation_string] = target_dir

        with open('files.txt', 'r') as f:
            for line in f:
                server_client_key, perturbation_string, source_file = line.strip().split(',')
                if server_client_key not in data_files:
                    data_files[server_client_key] = {}
                data_files[server_client_key][perturbation_string] = source_file

        return data_directories, data_files

    @staticmethod
    def setup_directory_structure_and_move_files(data_experiment, repertoire_data, servers, clients, perturbation_numbers_deplacement, protocole, codec, perturbation, nombre_hotes):
        current_directory = os.getcwd()
        repertoire_data_path = os.path.join(current_directory, repertoire_data)
        os.makedirs(repertoire_data_path, exist_ok=True)
        
        data_directories, data_files = DataOrganizer.read_data_from_files()

        # Supprimer le répertoire data_experiment après avoir déplacé tout son contenu
        #shutil.rmtree(os.path.join(current_directory, data_experiment), ignore_errors=True)
        
        print(f"Opération terminée : les répertoires et leur contenu ont été copiés vers '{repertoire_data_path}'.")

        # Impression des dictionnaires
        print("data_directories:", data_directories)
        print("data_files:", data_files)
        
        return data_directories, data_files


class Main:
    @staticmethod
    def main():
        setLogLevel('info')
        video_input = VideoInput()
        setup_script = SetupScript()
        #thread_experiments = []
        threads = []
        file_pattern = re.compile(r'.*\.ts$')
        #net = setup_script.setup_controller_connection()
        #net.start()
        #dumpNodeConnections(net.hosts)
        data_experiment = "data_experiment"
        repertoire_data = "repertoire_data"
        servers = video_input.get_servers()
        clients = video_input.get_clients()
        # ["1ms","2ms", "3ms","5ms","10ms","15ms"]
        HOST_NUMBERS = video_input.get_host_numbers()
        PERTURBATION_NUMBERS = video_input.get_perturbation_numbers()
        perturbation_numbers_deplacement = PERTURBATION_NUMBERS
        
        #ports = video_input.get_ports()
        #nginx_configs = video_input.get_nginx_configs()
        #video_urls = video_input.get_video_urls()
        #nom_interface_attendus = video_input.get_nom_interface_attendus()
        protocole = video_input.get_protocole()
        codec = video_input.get_codec()
        perturbation = video_input.get_perturbation()
        nombre_hotes = video_input.get_nombre_hotes()
        profile = video_input.get_profile()
        
###*#   # Begin part 2 experiment. 

        move_data_to_repertoire_data_bash_script_path = Main.trouver_fichier_par_extension("content_to_repertoire_bash_file", ".sh")
        
        # Déplacer le contenu de data_experiment vers  repertoire_data en exécutant le script bash "move_data_to_repertoire_data_bash_script_path"
        
        DataOrganizer.transfertContentDataExperimentToRepertoireData(
            move_data_to_repertoire_data_bash_script_path,
            data_experiment,
            repertoire_data,
            ','.join(servers),
            ','.join(clients),
            ','.join(map(str, perturbation_numbers_deplacement)),
            protocole,
            codec,
            perturbation,
            str(nombre_hotes)
        )

        # Configurer la structure des répertoires et déplacer les fichiers
        data_directories, data_files = DataOrganizer.setup_directory_structure_and_move_files(
            data_experiment, repertoire_data, servers, clients, perturbation_numbers_deplacement, protocole, codec, perturbation, nombre_hotes
        )
        bash_script_path = Main.trouver_fichier_par_extension("bash_file", ".sh")
        calculatorQoSmetrics = QoSMetricsCollector(data_files, bash_script_path)
        dict_data_qos_files_bitrate, dict_data_qos_files_packet_loss, dict_data_qos_files_average_latency, dict_data_qos_files_average_jitter = calculatorQoSmetrics.run()
        data_qos_dicts = {
            "bitrate": dict_data_qos_files_bitrate,
            "packet_loss": dict_data_qos_files_packet_loss,
            "average_latency": dict_data_qos_files_average_latency,
            "average_jitter": dict_data_qos_files_average_jitter
        }
        data_qos_names = {
            "bitrate": "dict_name_qos_bitrate",
            "packet_loss": "dict_name_qos_packetLoss",
            "average_latency": "dict_name_qos_averageLatency",
            "average_jitter": "dict_name_qos_averageJitter"
        }
        for key, value in data_qos_dicts.items():
            data_qos_resultats = data_qos_dicts[key]
            dict_name = data_qos_names[key]
            ExcelSaver.save_metrics_to_excel(data_qos_resultats, dict_name, codec, profile, perturbation, protocole, nombre_hotes)
        for actif_directory, perturbations_values in data_directories.items():
            for perturbation_value in perturbations_values:
                thread = threading.Thread(target=Main.process_directory2, args=(perturbation_value, file_pattern, actif_directory, protocole, codec, perturbation, nombre_hotes, profile))
                threads.append(thread)
                thread.start()
        for thread in threads:
            thread.join()
        data_video_files = Main.lister_videos_stream(repertoire_data, servers, clients, PERTURBATION_NUMBERS, protocole, codec, perturbation, nombre_hotes, profile)
        chemin_fichier_mp4 = Main.trouver_fichier_par_extension(f"file_video_serveur_{codec}", ".mp4")
        if chemin_fichier_mp4:
            print("Chemin absolu de la vidéo côté serveur:", chemin_fichier_mp4)
            calculator = QoEMetricsCollector(data_video_files, chemin_fichier_mp4)
            data_video_files_psnr, data_video_files_ssim = calculator.run()
        else:
            print("Aucun fichier .mp4 trouvé dans le répertoire courant.")
        data_qoe_dicts = {
            "psnr": data_video_files_psnr,
            "ssim": data_video_files_ssim
        }
        data_qoe_names = {
            "psnr": "dict_name_qoe_psnr",
            "ssim": "dict_name_qoe_ssim"
        }
        for key, value in data_qoe_dicts.items():
            data_qoe_resultats = data_qoe_dicts[key]
            dict_qoe_name = data_qoe_names[key]
            ExcelSaver.save_metrics_to_excel(data_qoe_resultats, dict_qoe_name, codec, profile, perturbation, protocole, nombre_hotes)
               
    @staticmethod
    def trouver_fichier_par_extension(directory, extension):
        repertoire_courant = os.path.abspath(os.getcwd())
        if directory == ".":
            repertoire_specifie = repertoire_courant
        else:
            if os.path.exists(directory):
                repertoire_specifie = os.path.abspath(directory)
            else:
                repertoire_specifie = os.path.abspath(os.path.join(repertoire_courant, directory))
        fichiers = os.listdir(repertoire_specifie)
        for fichier in fichiers:
            if fichier.endswith(extension):
                return os.path.abspath(os.path.join(repertoire_specifie, fichier))
        return None

    @staticmethod
    def process_directory2(perturbation_value, file_pattern, actif_directory, protocole, codec, perturbation, nombre_hotes, profile):
        base_directory = os.getcwd()
        current_directory = os.path.join(base_directory, "repertoire_data", f"{actif_directory}/{perturbation_value}/chunks-{actif_directory}_bbb_{codec}_{protocole}_hotes_{nombre_hotes}_{perturbation}_{perturbation_value}/")
        if not os.path.isdir(current_directory) or not os.listdir(current_directory):
            print(f"Le répertoire {current_directory} n'existe pas ou est vide.")
            return
        ts_files = sorted([os.path.join(current_directory, filename) for filename in os.listdir(current_directory) if file_pattern.match(filename)], key=Main.custom_sort)
        if not ts_files:
            print(f"Aucun fichier correspondant trouvé dans {current_directory}.")
            return
        playlist_filename = f"/tmp/ts_playlist_{actif_directory}_{perturbation_value}.txt"
        with open(playlist_filename, "w") as playlist_file:
            for file in ts_files:
                playlist_file.write(f"file '{file}'\n")
        output_video = os.path.join(base_directory, "repertoire_data", actif_directory, f"video_client_{actif_directory}_bbb_{codec}_{protocole}_hotes_{nombre_hotes}_{profile}_{perturbation}_{perturbation_value}.mp4")
        command = ["ffmpeg", "-f", "concat", "-safe", "0", "-i", playlist_filename, "-c", "copy", output_video]
        subprocess.run(command)
        if os.path.exists(output_video):
            print("La vidéo a été reconstituée avec succès.")
        else:
            print("Erreur lors de la reconstitution de la vidéo.")
        os.remove(playlist_filename)

    @staticmethod
    def custom_sort(filename):
        start = filename.rfind('k_') + 2
        end = filename.rfind('.ts')
        return int(filename[start:end])

    @staticmethod
    def lister_videos_stream(repertoire_data, servers, clients, PERTURBATION_NUMBERS_CREATION, protocole, codec, perturbation, nombre_hotes, profile):
        current_directory = os.getcwd()
        repertoire_data_path = os.path.join(current_directory, repertoire_data)
        os.makedirs(repertoire_data_path, exist_ok=True)
        data_video_files = {}
        for server, client in zip(servers, clients):
            server_client_key = f"{server}_{client}"
            sous_repertoire = os.path.join(repertoire_data_path, server_client_key)
            if os.path.exists(sous_repertoire) and os.path.isdir(sous_repertoire):
                if server_client_key not in data_video_files:
                    data_video_files[server_client_key] = {}
                for metrique_perturbation in PERTURBATION_NUMBERS_CREATION:
                    nom_fichier_cible = f"video_client_{server}_{client}_bbb_{codec}_{protocole}_hotes_{nombre_hotes}_{profile}_{perturbation}_{metrique_perturbation}.mp4"
                    chemin_fichier_cible = os.path.join(sous_repertoire, nom_fichier_cible)
                    if os.path.exists(chemin_fichier_cible):
                        data_video_files[server_client_key][metrique_perturbation] = chemin_fichier_cible
        return data_video_files

if __name__ == "__main__":
    Main.main()        




        
